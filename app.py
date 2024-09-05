#Read spreadsheet data
import openpyxl #to work with .xl files
import pyautogui #to work with your mouse and keyboard

workbook = openpyxl.load_workbook('vendas_de_produtos.xlsx') 
#it creates the variable to open the file
vendas_sheet = workbook['vendas']
#It creates another Var, that select wich page we are taking data

for linha in vendas_sheet.iter_rows(min_row=2):
    #e.g ["Murilo Barros", "Cadeira", "434", "Esportes"] iter_rows it will read all the lines starting in the row 2
    
    #using the mouse position (instructions bellow) you will catch the number that represents the field position from the system that u want to put your information(.value)
    
    #Input each cell from each line in a specific field of selected system
    
    #CLiente
    pyautogui.click(1140,408,duration=1.5) #.click receives the mouse position values and the time of execution
    pyautogui.write(linha[0].value) #.write receives "linha" value. 
    
    #Produto     
    pyautogui.click(1142,434,duration=1.5)
    pyautogui.write(linha[1].value)
    
    #Quantidade
    pyautogui.click(1158,459,duration=1.5)
    pyautogui.write(str(linha[2].value)) #pyautogui cant receive numbers(INT or Float), so you must convert to a string
    
    #Categoria
    pyautogui.click(1219,487,duration=1.5)
    pyautogui.write(linha[3].value)
    
    #Salvar e OK
    pyautogui.click(1095,516,duration=1.5)
    pyautogui.click(816,566,duration=1.5)
    
    
 #To catch the mouse position acess your powershell in windows 
# pip install mouseinfo (or pip3 to mac or linux)
# open python at powershell just typing python (or python3 to mac ou linux)
# from mouseinfo import mouseInfo (The second mouseInfo has the capital letter I)
#mouseInfo()